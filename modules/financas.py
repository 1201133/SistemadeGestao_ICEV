"""
Módulo de Gestão de Finanças
Páginas: Contas a Pagar | Extrato / Caixa | Resumo Financeiro
"""
import io
from datetime import date, datetime, timedelta
from typing import List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from database_financas import (
    adicionar_conta,
    adicionar_extrato,
    atualizar_conta,
    excluir_conta,
    excluir_extrato,
    listar_contas,
    listar_extrato,
)

# ─────────────────────────────────────────────────────────────────────────────
# Constantes de domínio
# ─────────────────────────────────────────────────────────────────────────────
STATUS_OPCOES = ["A Pagar", "Pago", "Parcelado"]
FORMA_PAGAMENTO_OPCOES = ["Boleto", "Depósito Bancário", "PIX", "Cheque", "Dinheiro"]
TIPO_EXTRATO_OPCOES = ["Crédito", "Débito"]

COR_VERDE = "1F7A4D"
COR_VERMELHO = "C0392B"
COR_AMARELO = "D4A017"
COR_AZUL_HEADER = "1F4E79"
COR_CINZA_LINHA = "F2F2F2"

# ─────────────────────────────────────────────────────────────────────────────
# Helpers de cálculo
# ─────────────────────────────────────────────────────────────────────────────

def _valor_parcela(valor: float, num_parcelas: int) -> float:
    """Valor de cada parcela (evita divisão por zero)."""
    return valor / num_parcelas if num_parcelas > 0 else valor


def _valor_pago(valor: float, num_parcelas: int, parcelas_pagas: int) -> float:
    return _valor_parcela(valor, num_parcelas) * parcelas_pagas


def _saldo_devedor(valor: float, num_parcelas: int, parcelas_pagas: int, status: str) -> float:
    if status == "Pago":
        return 0.0
    return valor - _valor_pago(valor, num_parcelas, parcelas_pagas)


def _situacao_vencimento(data_venc_str: str, status: str) -> str:
    if status == "Pago":
        return "✅ Pago"
    try:
        data_venc = datetime.strptime(data_venc_str, "%Y-%m-%d").date()
    except ValueError:
        return "-"
    hoje = date.today()
    dias = (data_venc - hoje).days
    if dias < 0:
        return f"🔴 Vencido há {abs(dias)} dia(s)"
    if dias == 0:
        return "🟠 Vence hoje"
    if dias <= 7:
        return f"🟡 Vence em {dias} dia(s)"
    return f"🟢 Vence em {dias} dia(s)"


def _calcular_totais_contas(contas: List[Tuple]) -> dict:
    """Retorna um dicionário com os totais financeiros das contas."""
    total_geral = sum(c[3] for c in contas)
    total_pago = sum(c[3] for c in contas if c[6] == "Pago")
    total_a_pagar = sum(
        _saldo_devedor(c[3], c[8], c[9], c[6])
        for c in contas if c[6] != "Pago"
    )
    hoje = date.today()
    total_vencido = 0.0
    for c in contas:
        if c[6] == "Pago":
            continue
        data_venc = _str_to_date(c[2])
        if data_venc is not None and data_venc < hoje:
            total_vencido += _saldo_devedor(c[3], c[8], c[9], c[6])
    return {
        "total_geral": total_geral,
        "total_pago": total_pago,
        "total_a_pagar": total_a_pagar,
        "total_vencido": total_vencido,
    }


def _calcular_saldo_banco(extrato: List[Tuple]) -> float:
    """Saldo bancário = Créditos - Débitos."""
    saldo = 0.0
    for e in extrato:
        # e[3] = valor, e[4] = tipo
        if e[4] == "Crédito":
            saldo += e[3]
        else:
            saldo -= e[3]
    return saldo


def _str_to_date(s: str) -> Optional[date]:
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def _fmt_data_br(data_str: str) -> str:
    """Formata data ISO (YYYY-MM-DD) para padrão brasileiro (DD/MM/YYYY)."""
    try:
        return datetime.strptime(data_str, "%Y-%m-%d").strftime("%d/%m/%Y")
    except (ValueError, TypeError):
        return data_str or ""


def _fmt_brl(valor: float) -> str:
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


# ─────────────────────────────────────────────────────────────────────────────
# DataFrame helpers
# ─────────────────────────────────────────────────────────────────────────────

def _contas_to_df(contas: List[Tuple]) -> pd.DataFrame:
    """Converte lista de tuplas em DataFrame enriquecido com colunas calculadas."""
    linhas = []
    for c in contas:
        (
            id_, data_lanc, data_venc, valor, motivo, fornecedor,
            status, forma_pag, num_parcelas, parcelas_pagas, usuario, _
        ) = c

        vp = _valor_parcela(valor, num_parcelas)
        v_pago = _valor_pago(valor, num_parcelas, parcelas_pagas)
        saldo = _saldo_devedor(valor, num_parcelas, parcelas_pagas, status)
        situacao = _situacao_vencimento(data_venc, status)

        linhas.append({
            "ID": id_,
            "Data Lançamento": _fmt_data_br(data_lanc),
            "Data Vencimento": _fmt_data_br(data_venc),
            "Fornecedor": fornecedor,
            "Motivo / Ocorrência": motivo,
            "Valor Total (R$)": valor,
            "Status": status,
            "Forma Pgto": forma_pag,
            "Nº Parcelas": num_parcelas,
            "Parcelas Pagas": parcelas_pagas,
            "Valor Parcela (R$)": round(vp, 2),
            "Valor Pago (R$)": round(v_pago, 2),
            "Saldo Devedor (R$)": round(saldo, 2),
            "Situação Vencimento": situacao,
            "Usuário": usuario,
        })
    return pd.DataFrame(linhas)


def _extrato_to_df(extrato: List[Tuple]) -> pd.DataFrame:
    linhas = []
    saldo_acumulado = 0.0
    for e in reversed(extrato):  # ordem cronológica
        (id_, data, descricao, valor, tipo, usuario, _) = e
        if tipo == "Crédito":
            saldo_acumulado += valor
        else:
            saldo_acumulado -= valor
        linhas.append({
            "ID": id_,
            "Data": _fmt_data_br(data),
            "Descrição": descricao,
            "Tipo": tipo,
            "Valor (R$)": valor,
            "Saldo Acumulado (R$)": round(saldo_acumulado, 2),
            "Usuário": usuario,
        })
    return pd.DataFrame(list(reversed(linhas)))


# ─────────────────────────────────────────────────────────────────────────────
# Exportação Excel
# ─────────────────────────────────────────────────────────────────────────────

def _estilo_header(ws, row: int, col_start: int, col_end: int, cor_hex: str = COR_AZUL_HEADER):
    fill = PatternFill("solid", fgColor=cor_hex)
    font = Font(bold=True, color="FFFFFF", size=10)
    aln  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = aln
        cell.border = borda


def _aplicar_borda_linha(ws, row: int, col_start: int, col_end: int, zebra: bool = False):
    fill = PatternFill("solid", fgColor=COR_CINZA_LINHA) if zebra else PatternFill()
    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = borda
        if zebra:
            cell.fill = fill
        cell.alignment = Alignment(vertical="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def gerar_excel(contas: List[Tuple], extrato: List[Tuple]) -> bytes:
    """
    Gera um workbook Excel com três abas:
    1. Contas a Pagar  — todos os lançamentos com colunas calculadas
    2. Extrato / Caixa — histórico bancário com saldo acumulado
    3. Painel Financeiro — resumo executivo com lógica de caixa
    """
    wb = Workbook()

    # ── Aba 1: Contas a Pagar ────────────────────────────────────────────────
    ws1 = wb.active
    if ws1 is None:
        raise RuntimeError("Falha ao criar planilha: aba ativa indisponivel.")
    ws1.title = "Contas a Pagar"

    cabecalho = [
        "ID", "Data Lançamento", "Data Vencimento", "Fornecedor",
        "Motivo / Ocorrência", "Valor Total (R$)", "Status",
        "Forma de Pagamento", "Nº Parcelas", "Parcelas Pagas",
        "Valor Parcela (R$)", "Valor Pago (R$)", "Saldo Devedor (R$)",
        "Dias p/ Vencimento", "Usuário",
    ]
    ws1.append(cabecalho)
    _estilo_header(ws1, 1, 1, len(cabecalho))
    ws1.row_dimensions[1].height = 30

    hoje = date.today()
    for i, c in enumerate(contas, start=2):
        (
            id_, data_lanc, data_venc, valor, motivo, fornecedor,
            status, forma_pag, num_parcelas, parcelas_pagas, usuario, _
        ) = c
        vp    = round(_valor_parcela(valor, num_parcelas), 2)
        v_pago = round(_valor_pago(valor, num_parcelas, parcelas_pagas), 2)
        saldo  = round(_saldo_devedor(valor, num_parcelas, parcelas_pagas, status), 2)
        data_v = _str_to_date(data_venc)
        dias   = (data_v - hoje).days if data_v else ""

        ws1.append([
            id_, _fmt_data_br(data_lanc), _fmt_data_br(data_venc), fornecedor, motivo,
            valor, status, forma_pag, num_parcelas, parcelas_pagas,
            vp, v_pago, saldo, dias, usuario,
        ])
        zebra = (i % 2 == 0)
        _aplicar_borda_linha(ws1, i, 1, len(cabecalho), zebra)

        # Colorir célula de status
        cor_status = {
            "Pago": COR_VERDE,
            "A Pagar": COR_AMARELO,
            "Parcelado": "2E75B6",
        }.get(status, "000000")
        cell_status = ws1.cell(row=i, column=7)
        cell_status.font = Font(bold=True, color=cor_status)

        # Colorir saldo devedor
        cell_saldo = ws1.cell(row=i, column=13)
        cell_saldo.font = Font(color=COR_VERMELHO if saldo > 0 else COR_VERDE, bold=True)

    # Linha de totais
    n_col = len(cabecalho)
    linha_total = ws1.max_row + 1
    ws1.cell(row=linha_total, column=4, value="TOTAIS")
    ws1.cell(row=linha_total, column=4).font = Font(bold=True)
    ws1.cell(row=linha_total, column=6, value=sum(c[3] for c in contas))
    ws1.cell(row=linha_total, column=6).font = Font(bold=True)
    ws1.cell(row=linha_total, column=12, value=sum(
        _valor_pago(c[3], c[8], c[9]) for c in contas
    ))
    ws1.cell(row=linha_total, column=13, value=round(sum(
        _saldo_devedor(c[3], c[8], c[9], c[6]) for c in contas
    ), 2))
    ws1.cell(row=linha_total, column=13).font = Font(bold=True, color=COR_VERMELHO)
    _estilo_header(ws1, linha_total, 1, n_col, "2D4A6E")

    _auto_width(ws1)

    # ── Aba 2: Extrato / Caixa ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Extrato - Caixa")
    cab2 = ["ID", "Data", "Descrição", "Tipo", "Valor (R$)", "Saldo Acumulado (R$)", "Usuário"]
    ws2.append(cab2)
    _estilo_header(ws2, 1, 1, len(cab2), "1F4E79")
    ws2.row_dimensions[1].height = 30

    saldo_acum = 0.0
    for i, e in enumerate(reversed(extrato), start=2):
        (id_, data, descricao, valor, tipo, usuario, _) = e
        if tipo == "Crédito":
            saldo_acum += valor
        else:
            saldo_acum -= valor
        ws2.append([id_, _fmt_data_br(data), descricao, tipo, valor, round(saldo_acum, 2), usuario])
        zebra = (i % 2 == 0)
        _aplicar_borda_linha(ws2, i, 1, len(cab2), zebra)
        cor_tipo = COR_VERDE if tipo == "Crédito" else COR_VERMELHO
        ws2.cell(row=i, column=4).font = Font(color=cor_tipo, bold=True)
        ws2.cell(row=i, column=6).font = Font(
            color=COR_VERDE if saldo_acum >= 0 else COR_VERMELHO, bold=True
        )

    _auto_width(ws2)

    # ── Aba 3: Painel Financeiro ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Painel Financeiro")

    def _escrever_bloco(ws, linha_ini: int, titulo: str, itens: list[tuple], cor: str):
        """Escreve um bloco de título + linhas chave/valor."""
        ws.merge_cells(start_row=linha_ini, start_column=1, end_row=linha_ini, end_column=3)
        cell_titulo = ws.cell(row=linha_ini, column=1, value=titulo)
        cell_titulo.fill = PatternFill("solid", fgColor=cor)
        cell_titulo.font = Font(bold=True, color="FFFFFF", size=11)
        cell_titulo.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[linha_ini].height = 22

        for j, (chave, valor) in enumerate(itens, start=linha_ini + 1):
            c1 = ws.cell(row=j, column=1, value=chave)
            c2 = ws.cell(row=j, column=2, value=valor)
            c1.font = Font(bold=True)
            c2.alignment = Alignment(horizontal="right")
            _aplicar_borda_linha(ws, j, 1, 3, j % 2 == 0)

        return linha_ini + len(itens) + 2  # próxima linha disponível

    totais = _calcular_totais_contas(contas)
    saldo_banco = _calcular_saldo_banco(extrato)
    saldo_disponivel = saldo_banco - totais["total_a_pagar"]
    total_creditos = sum(e[3] for e in extrato if e[4] == "Crédito")
    total_debitos  = sum(e[3] for e in extrato if e[4] == "Débito")

    ws3.column_dimensions["A"].width = 34
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 5

    linha = 1
    ws3.cell(row=linha, column=1, value="PAINEL FINANCEIRO — GERADO EM " + datetime.now().strftime("%d/%m/%Y %H:%M"))
    ws3.cell(row=linha, column=1).font = Font(bold=True, size=13, color=COR_AZUL_HEADER)
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws3.row_dimensions[1].height = 28
    linha = 3

    linha = _escrever_bloco(ws3, linha, "💳 CONTAS A PAGAR", [
        ("Total de lançamentos (R$)",    f"{totais['total_geral']:,.2f}"),
        ("Total já pago (R$)",           f"{totais['total_pago']:,.2f}"),
        ("Total pendente / em aberto (R$)", f"{totais['total_a_pagar']:,.2f}"),
        ("Total vencido e não pago (R$)", f"{totais['total_vencido']:,.2f}"),
    ], COR_AZUL_HEADER)

    linha = _escrever_bloco(ws3, linha, "🏦 EXTRATO BANCÁRIO", [
        ("Total de créditos recebidos (R$)", f"{total_creditos:,.2f}"),
        ("Total de débitos registrados (R$)", f"{total_debitos:,.2f}"),
        ("Saldo atual em caixa/conta (R$)",   f"{saldo_banco:,.2f}"),
    ], "1F7A4D")

    linha = _escrever_bloco(ws3, linha, "📊 PROJEÇÃO DE CAIXA", [
        ("Saldo em conta (R$)",               f"{saldo_banco:,.2f}"),
        ("(−) Total de contas a pagar (R$)",  f"{totais['total_a_pagar']:,.2f}"),
        ("(=) Saldo disponível projetado (R$)", f"{saldo_disponivel:,.2f}"),
        ("⚠️  Situação", "POSITIVO ✅" if saldo_disponivel >= 0 else "DÉFICIT 🔴"),
    ], COR_AMARELO if saldo_disponivel < 0 else "1F7A4D")

    linha = _escrever_bloco(ws3, linha, "ℹ️  LÓGICA DE CÁLCULO", [
        ("Valor da Parcela",        "= Valor Total ÷ Nº de Parcelas"),
        ("Valor Pago",              "= Valor da Parcela × Parcelas Pagas"),
        ("Saldo Devedor",           "= Valor Total − Valor Pago"),
        ("Saldo em Caixa",          "= Σ Créditos − Σ Débitos (extrato)"),
        ("Saldo Disponível Proj.",  "= Saldo em Caixa − Total Pendente"),
    ], "555555")

    _auto_width(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# UI — Aba Contas a Pagar
# ─────────────────────────────────────────────────────────────────────────────

def _aba_contas_a_pagar(usuario: str, nivel: str):
    st.markdown("### 📋 Contas a Pagar")

    contas = listar_contas(usuario, nivel)
    totais = _calcular_totais_contas(contas) if contas else {}

    # KPIs rápidos
    if totais:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total Geral", _fmt_brl(totais["total_geral"]))
        c2.metric("Total Pago", _fmt_brl(totais["total_pago"]), delta="pago")
        c3.metric("Total em Aberto", _fmt_brl(totais["total_a_pagar"]),
                  delta=f"−{_fmt_brl(totais['total_a_pagar'])}", delta_color="inverse")
        c4.metric("Total Vencido", _fmt_brl(totais["total_vencido"]),
                  delta="vencido" if totais["total_vencido"] > 0 else "ok",
                  delta_color="inverse" if totais["total_vencido"] > 0 else "normal")

    st.markdown("---")

    with st.expander("➕ Nova Conta a Pagar", expanded=not contas):
        with st.form("form_nova_conta", clear_on_submit=True):
            col1, col2 = st.columns(2)
            with col1:
                data_lanc = st.date_input(
                    "📅 Data do Lançamento",
                    value=date.today(),
                    format="DD/MM/YYYY",
                )
                valor = st.number_input("💲 Valor Total (R$)", min_value=0.01, step=0.01, format="%.2f")
                motivo = st.text_input("📝 Motivo / Ocorrência", placeholder="Ex: Conta de luz Maio/2026")
                fornecedor = st.text_input("🏢 Fornecedor", placeholder="Ex: ENERGISA")
            with col2:
                data_venc = st.date_input(
                    "⏰ Data de Vencimento",
                    value=date.today() + timedelta(days=30),
                    format="DD/MM/YYYY",
                )
                status = st.selectbox("📌 Status", STATUS_OPCOES)
                forma_pag = st.selectbox("💳 Forma de Pagamento", FORMA_PAGAMENTO_OPCOES)

            col3, col4 = st.columns(2)
            with col3:
                st.markdown("🔢 Número de Parcelas")
                num_parcelas = st.number_input(
                    "Número de Parcelas",
                    min_value=1, max_value=360, value=1, step=1,
                    help="Para pagamento único, deixe 1",
                    label_visibility="collapsed",
                )
            with col4:
                st.markdown("✅ Parcelas Já Pagas")
                parcelas_pagas = st.number_input(
                    "Parcelas Já Pagas",
                    min_value=0, max_value=360, value=0, step=1,
                    label_visibility="collapsed",
                )

            # Cálculo em tempo real (preview antes de salvar)
            if valor and num_parcelas:
                vp_prev = valor / num_parcelas
                v_pago_prev = vp_prev * parcelas_pagas
                saldo_prev = valor - v_pago_prev
                st.info(
                    f"**Prévia dos cálculos:** "
                    f"Parcela = {_fmt_brl(vp_prev)} | "
                    f"Pago = {_fmt_brl(v_pago_prev)} | "
                    f"Saldo Devedor = {_fmt_brl(saldo_prev)}"
                )

            salvar = st.form_submit_button("💾 Salvar Conta", type="primary")
            if salvar:
                erros = []
                if not motivo.strip():
                    erros.append("Motivo / Ocorrência é obrigatório.")
                if not fornecedor.strip():
                    erros.append("Fornecedor é obrigatório.")
                if parcelas_pagas > num_parcelas:
                    erros.append("Parcelas pagas não podem ser maiores que o número de parcelas.")
                if erros:
                    for e in erros:
                        st.error(e)
                else:
                    ok = adicionar_conta(
                        data_lancamento=data_lanc.strftime("%Y-%m-%d"),
                        data_vencimento=data_venc.strftime("%Y-%m-%d"),
                        valor=float(valor),
                        motivo=motivo.strip(),
                        fornecedor=fornecedor.strip(),
                        status=status,
                        forma_pagamento=forma_pag,
                        num_parcelas=int(num_parcelas),
                        parcelas_pagas=int(parcelas_pagas),
                        usuario=usuario,
                    )
                    if ok:
                        st.success("✅ Conta registrada com sucesso!")
                        st.rerun()
                    else:
                        st.error("❌ Erro ao salvar. Tente novamente.")

    # Tabela de contas existentes
    st.markdown("---")
    st.markdown("#### 📊 Lançamentos Registrados")

    if not contas:
        st.info("ℹ️ Nenhuma conta cadastrada ainda.")
        return

    # Filtros
    cf1, cf2 = st.columns(2)
    filtro_status = cf1.multiselect(
        "Filtrar por Status", STATUS_OPCOES,
        default=STATUS_OPCOES,
        key="filtro_status_contas",
    )
    filtro_fornecedor = cf2.text_input("Filtrar por Fornecedor", key="filtro_forn")

    contas_filtradas = [
        c for c in contas
        if c[6] in filtro_status
        and (not filtro_fornecedor or filtro_fornecedor.lower() in c[5].lower())
    ]

    if not contas_filtradas:
        st.warning("Nenhum registro encontrado para os filtros aplicados.")
        return

    df = _contas_to_df(contas_filtradas)
    st.dataframe(df, use_container_width=True, hide_index=True)

    # Edição / exclusão
    st.markdown("---")
    with st.expander("✏️ Editar ou Excluir uma Conta"):
        ids_disponiveis = [c[0] for c in contas_filtradas]
        conta_id_sel = st.selectbox("Selecione o ID da conta", ids_disponiveis, key="sel_conta_edit")

        conta_sel = next((c for c in contas_filtradas if c[0] == conta_id_sel), None)
        if conta_sel:
            (
                _id, data_lanc_s, data_venc_s, valor_s, motivo_s, forn_s,
                status_s, forma_s, n_parc_s, pagas_s, usr_s, _
            ) = conta_sel

            with st.form("form_editar_conta"):
                ec1, ec2 = st.columns(2)
                with ec1:
                    e_data_lanc = st.date_input(
                        "Data Lançamento",
                        value=_str_to_date(data_lanc_s) or date.today(),
                        key="e_dl",
                        format="DD/MM/YYYY",
                    )
                    e_valor = st.number_input("Valor Total (R$)", value=float(valor_s),
                        min_value=0.01, step=0.01, format="%.2f", key="e_val")
                    e_motivo = st.text_input("Motivo / Ocorrência", value=motivo_s, key="e_mot")
                    e_forn = st.text_input("Fornecedor", value=forn_s, key="e_forn")
                with ec2:
                    e_data_venc = st.date_input(
                        "Data Vencimento",
                        value=_str_to_date(data_venc_s) or date.today(),
                        key="e_dv",
                        format="DD/MM/YYYY",
                    )
                    e_status = st.selectbox("Status", STATUS_OPCOES,
                        index=STATUS_OPCOES.index(status_s) if status_s in STATUS_OPCOES else 0,
                        key="e_status")
                    e_forma = st.selectbox("Forma de Pagamento", FORMA_PAGAMENTO_OPCOES,
                        index=FORMA_PAGAMENTO_OPCOES.index(forma_s) if forma_s in FORMA_PAGAMENTO_OPCOES else 0,
                        key="e_forma")
                ec3, ec4 = st.columns(2)
                with ec3:
                    st.markdown("🔢 Nº Parcelas")
                    e_nparc = st.number_input("Nº Parcelas", value=int(n_parc_s),
                        min_value=1, max_value=360, step=1, key="e_np", label_visibility="collapsed")
                with ec4:
                    st.markdown("✅ Parcelas Pagas")
                    e_pagas = st.number_input("Parcelas Pagas", value=int(pagas_s),
                        min_value=0, max_value=360, step=1, key="e_pg", label_visibility="collapsed")

                col_ed1, col_ed2 = st.columns(2)
                with col_ed1:
                    salvar_edicao = st.form_submit_button("💾 Salvar Alterações", type="primary")
                with col_ed2:
                    excluir_btn = st.form_submit_button("🗑️ Excluir Conta", type="secondary")

                if salvar_edicao:
                    motivo_limpo = (e_motivo or "").strip()
                    fornecedor_limpo = (e_forn or "").strip()
                    if not motivo_limpo or not fornecedor_limpo:
                        st.error("Motivo e fornecedor sao obrigatorios para atualizar a conta.")
                    else:
                        ok = atualizar_conta(
                            conta_id_sel,
                            data_lancamento=e_data_lanc.strftime("%Y-%m-%d"),
                            data_vencimento=e_data_venc.strftime("%Y-%m-%d"),
                            valor=float(e_valor),
                            motivo=motivo_limpo,
                            fornecedor=fornecedor_limpo,
                            status=e_status,
                            forma_pagamento=e_forma,
                            num_parcelas=int(e_nparc),
                            parcelas_pagas=int(e_pagas),
                        )
                        if ok:
                            st.success("✅ Conta atualizada!")
                            st.rerun()
                        else:
                            st.error("❌ Erro ao atualizar.")

                if excluir_btn:
                    ok = excluir_conta(conta_id_sel)
                    if ok:
                        st.success("✅ Conta excluída!")
                        st.rerun()
                    else:
                        st.error("❌ Erro ao excluir.")


# ─────────────────────────────────────────────────────────────────────────────
# UI — Aba Extrato / Caixa
# ─────────────────────────────────────────────────────────────────────────────

def _aba_extrato_caixa(usuario: str, nivel: str):
    st.markdown("### 🏦 Extrato / Caixa")

    extrato = listar_extrato(usuario, nivel)
    saldo_atual = _calcular_saldo_banco(extrato)
    total_cred  = sum(e[3] for e in extrato if e[4] == "Crédito")
    total_deb   = sum(e[3] for e in extrato if e[4] == "Débito")

    m1, m2, m3 = st.columns(3)
    m1.metric("Total Créditos", _fmt_brl(total_cred))
    m2.metric("Total Débitos", _fmt_brl(total_deb))
    m3.metric(
        "Saldo em Caixa",
        _fmt_brl(saldo_atual),
        delta="positivo" if saldo_atual >= 0 else "negativo",
        delta_color="normal" if saldo_atual >= 0 else "inverse",
    )

    st.markdown("---")
    with st.expander("➕ Adicionar Lançamento Bancário", expanded=not extrato):
        with st.form("form_extrato", clear_on_submit=True):
            col1, col2 = st.columns(2)
            with col1:
                ex_data = st.date_input(
                    "Data",
                    value=date.today(),
                    key="ex_data",
                    format="DD/MM/YYYY",
                )
                ex_desc = st.text_input("Descrição", placeholder="Ex: Depósito de dízimos", key="ex_desc")
            with col2:
                ex_tipo = st.selectbox("Tipo", TIPO_EXTRATO_OPCOES, key="ex_tipo")
                ex_valor = st.number_input("Valor (R$)", min_value=0.01, step=0.01,
                    format="%.2f", key="ex_valor")

            salvar_ext = st.form_submit_button("💾 Salvar Lançamento", type="primary")
            if salvar_ext:
                if not ex_desc.strip():
                    st.error("Descrição é obrigatória.")
                else:
                    ok = adicionar_extrato(
                        data=ex_data.strftime("%Y-%m-%d"),
                        descricao=ex_desc.strip(),
                        valor=float(ex_valor),
                        tipo=ex_tipo,
                        usuario=usuario,
                    )
                    if ok:
                        st.success("✅ Lançamento adicionado!")
                        st.rerun()
                    else:
                        st.error("❌ Erro ao salvar.")

    st.markdown("---")
    st.markdown("#### 📋 Histórico do Extrato")

    if not extrato:
        st.info("ℹ️ Nenhum lançamento bancário registrado ainda.")
        return

    df_ext = _extrato_to_df(extrato)
    st.dataframe(df_ext, use_container_width=True, hide_index=True)

    with st.expander("🗑️ Excluir Lançamento do Extrato"):
        ids_ext = [e[0] for e in extrato]
        sel_ext_id = st.selectbox("Selecione o ID", ids_ext, key="sel_ext_del")
        if st.button("Excluir", key="btn_del_ext"):
            ok = excluir_extrato(sel_ext_id)
            if ok:
                st.success("✅ Lançamento excluído!")
                st.rerun()
            else:
                st.error("❌ Erro ao excluir.")


# ─────────────────────────────────────────────────────────────────────────────
# UI — Aba Resumo Financeiro
# ─────────────────────────────────────────────────────────────────────────────

def _aba_resumo(usuario: str, nivel: str):
    st.markdown("### 📊 Resumo Financeiro")

    contas  = listar_contas(usuario, nivel)
    extrato = listar_extrato(usuario, nivel)

    totais       = _calcular_totais_contas(contas) if contas else {
        "total_geral": 0, "total_pago": 0,
        "total_a_pagar": 0, "total_vencido": 0,
    }
    saldo_banco  = _calcular_saldo_banco(extrato)
    saldo_disp   = saldo_banco - totais["total_a_pagar"]

    st.markdown("#### 💳 Contas a Pagar")
    ca1, ca2, ca3, ca4 = st.columns(4)
    ca1.metric("Total Registrado", _fmt_brl(totais["total_geral"]))
    ca2.metric("Pago", _fmt_brl(totais["total_pago"]))
    ca3.metric("Em Aberto", _fmt_brl(totais["total_a_pagar"]))
    ca4.metric("Vencido", _fmt_brl(totais["total_vencido"]))

    st.markdown("#### 🏦 Caixa / Banco")
    cb1, cb2, cb3 = st.columns(3)
    total_cred = sum(e[3] for e in extrato if e[4] == "Crédito")
    total_deb  = sum(e[3] for e in extrato if e[4] == "Débito")
    cb1.metric("Créditos", _fmt_brl(total_cred))
    cb2.metric("Débitos", _fmt_brl(total_deb))
    cb3.metric("Saldo em Caixa", _fmt_brl(saldo_banco))

    st.markdown("#### 📈 Projeção de Caixa")
    cor_sinal = "normal" if saldo_disp >= 0 else "inverse"
    st.metric(
        label="Saldo Disponível Projetado (Caixa − Contas em Aberto)",
        value=_fmt_brl(saldo_disp),
        delta="Positivo ✅" if saldo_disp >= 0 else "Déficit 🔴",
        delta_color=cor_sinal,
    )

    st.markdown("---")
    st.markdown("""
    **🔢 Lógica dos Cálculos**

    | Campo | Fórmula |
    |---|---|
    | Valor da Parcela | Valor Total ÷ Nº de Parcelas |
    | Valor Pago | Valor da Parcela × Parcelas Pagas |
    | Saldo Devedor | Valor Total − Valor Pago |
    | Saldo em Caixa | Σ Créditos − Σ Débitos (extrato bancário) |
    | Saldo Disponível Projetado | Saldo em Caixa − Total de Contas em Aberto |
    """)

    # Exemplo de lógica para o usuário verificar
    with st.expander("📌 Exemplo de Lógica — Demonstração"):
        exemplo = pd.DataFrame({
            "Fornecedor":    ["Energisa", "Internet",  "Aluguel",   "Material"],
            "Valor Total":   [350.00,     199.90,      1200.00,     450.00],
            "Status":        ["Pago",     "A Pagar",   "Parcelado", "A Pagar"],
            "Nº Parcelas":   [1,          1,           12,          1],
            "Parcelas Pagas":[1,          0,           3,           0],
        })
        exemplo["Valor Parcela"]   = exemplo["Valor Total"] / exemplo["Nº Parcelas"]
        exemplo["Valor Pago"]      = exemplo["Valor Parcela"] * exemplo["Parcelas Pagas"]
        exemplo["Saldo Devedor"]   = exemplo.apply(
            lambda r: 0.0 if r["Status"] == "Pago"
            else r["Valor Total"] - r["Valor Pago"], axis=1
        )
        st.dataframe(exemplo, use_container_width=True, hide_index=True)

        saldo_ex_banco = 2500.00
        total_aberto_ex = exemplo["Saldo Devedor"].sum()
        saldo_proj_ex = saldo_ex_banco - total_aberto_ex
        st.markdown(f"""
        **Extrato bancário (exemplo):** Saldo em conta = **R$ 2.500,00**

        | | Valor |
        |---|---|
        | Saldo em caixa | R$ 2.500,00 |
        | (−) Total em aberto | R$ {total_aberto_ex:,.2f} |
        | **(=) Saldo Disponível Projetado** | **R$ {saldo_proj_ex:,.2f}** |

        {'✅ Situação: **POSITIVO** — há saldo suficiente para cobrir todas as despesas.' if saldo_proj_ex >= 0 else '🔴 Situação: **DÉFICIT** — o saldo em caixa é insuficiente para as despesas em aberto.'}
        """)

    # Botão de download Excel
    st.markdown("---")
    st.markdown("#### ⬇️ Exportar Planilha")
    if st.button("📥 Gerar e Baixar Planilha Excel", type="primary", key="btn_download_excel"):
        contas_all  = listar_contas(usuario, nivel)
        extrato_all = listar_extrato(usuario, nivel)
        excel_bytes = gerar_excel(contas_all, extrato_all)
        nome_arquivo = f"gestao_financas_{datetime.now().strftime('%d-%m-%Y_%H%M')}.xlsx"
        st.download_button(
            label="📊 Clique aqui para baixar",
            data=excel_bytes,
            file_name=nome_arquivo,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_excel_real",
        )


# ─────────────────────────────────────────────────────────────────────────────
# Ponto de entrada do módulo
# ─────────────────────────────────────────────────────────────────────────────

def exibir_pagina_financas():
    """Página principal do módulo Gestão de Finanças."""
    st.subheader("💰 Gestão de Finanças")

    usuario = st.session_state["usuario"]
    nivel   = st.session_state["nivel"]

    aba1, aba2, aba3 = st.tabs([
        "📋 Contas a Pagar",
        "🏦 Extrato / Caixa",
        "📊 Resumo Financeiro",
    ])

    with aba1:
        _aba_contas_a_pagar(usuario, nivel)

    with aba2:
        _aba_extrato_caixa(usuario, nivel)

    with aba3:
        _aba_resumo(usuario, nivel)
